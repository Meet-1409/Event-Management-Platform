from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.conf import settings
from django.utils import timezone
from django.core.mail import send_mail
from django.template.loader import render_to_string
import stripe
import json
import logging
from decimal import Decimal
from .models import Payment, PaymentMethod, PaymentGateway, Invoice, Refund
from events.models import Event
from events.utils import generate_pdf_invoice
import qrcode
import base64
from io import BytesIO
from django.utils import timezone
from datetime import timedelta
from django.db.models import Q

# Configure Stripe - with dev mode safety
if hasattr(settings, 'STRIPE_SECRET_KEY') and settings.STRIPE_SECRET_KEY.startswith('sk_'):
    stripe.api_key = settings.STRIPE_SECRET_KEY
else:
    stripe.api_key = None

logger = logging.getLogger(__name__)

@login_required
def payment_list(request):
    """Display user's payment history"""
    payments = Payment.objects.filter(user=request.user).order_by('-created_at')
    
    context = {
        'payments': payments,
        'total_paid': sum(payment.amount for payment in payments if payment.status == 'completed'),
        'pending_payments': payments.filter(status='pending').count(),
    }
    
    return render(request, 'payments/payment_list.html', context)

@login_required
def process_payment(request):
    """Process payment for event booking"""
    if request.method == 'POST':
        try:
            # Get payment data
            event_id = request.POST.get('event_id')
            amount = Decimal(request.POST.get('amount'))
            payment_method = request.POST.get('payment_method')
            currency = request.POST.get('currency', 'INR')
            
            # Get event
            event = get_object_or_404(Event, id=event_id)
            
            # Create invoice
            invoice = Invoice.objects.create(
                event=event,
                user=request.user,
                issue_date=timezone.now().date(),
                due_date=(timezone.now() + timezone.timedelta(days=30)).date(),
                subtotal=amount,
                tax_amount=amount * Decimal('0.18'),  # 18% GST
                total_amount=amount * Decimal('1.18'),
                currency=currency,
                status='sent'
            )
            
            # Get payment method
            payment_method_obj = get_object_or_404(PaymentMethod, name=payment_method)
            
            # Create payment record
            payment = Payment.objects.create(
                invoice=invoice,
                user=request.user,
                amount=invoice.total_amount,
                currency=currency,
                payment_method=payment_method_obj,
                status='pending'
            )
            
            # Check if in dev mode
            if getattr(settings, 'PAYMENT_DEV_MODE', False):
                # Dev mode - mock successful payment
                payment.status = 'completed'
                payment.payment_date = timezone.now()
                payment.save()
                
                # Update invoice
                invoice.status = 'paid'
                invoice.paid_amount = payment.amount
                invoice.save()
                
                return JsonResponse({
                    'success': True,
                    'dev_mode': True,
                    'message': 'DEV MODE: Payment simulated successfully',
                    'payment_id': payment.payment_id
                })
            
            if payment_method == 'stripe':
                # Process with Stripe
                return process_stripe_payment(request, payment)
            elif payment_method == 'razorpay':
                # Process with Razorpay
                return process_razorpay_payment(request, payment)
            else:
                # Process with other methods
                return process_other_payment(request, payment)
                
        except Exception as e:
            logger.error(f"Payment processing error: {str(e)}")
            messages.error(request, 'Payment processing failed. Please try again.')
            return JsonResponse({'success': False, 'error': str(e)})
    
    # GET request - show payment form
    event_id = request.GET.get('event_id')
    if event_id:
        event = get_object_or_404(Event, id=event_id)
        
        # Calculate amounts
        base_amount = event.total_budget
        service_fee = base_amount * Decimal('0.05')  # 5% service fee
        gst_amount = (base_amount + service_fee) * Decimal('0.18')  # 18% GST
        total_amount = base_amount + service_fee + gst_amount
        
        context = {
            'event': event,
            'base_amount': base_amount,
            'service_fee': service_fee,
            'gst_amount': gst_amount,
            'total_amount': total_amount,
            'stripe_publishable_key': settings.STRIPE_PUBLISHABLE_KEY,
            'razorpay_key_id': getattr(settings, 'RAZORPAY_KEY_ID', ''),
            'today': timezone.now().date(),
        }
        
        return render(request, 'payments/process_payment.html', context)
    
    messages.error(request, 'No event specified for payment.')
    return redirect('events:event_list')

def process_stripe_payment(request, payment):
    """Process payment using Stripe"""
    try:
        # Create Stripe payment intent
        intent = stripe.PaymentIntent.create(
            amount=int(payment.amount * 100),  # Convert to cents
            currency=payment.currency.lower(),
            metadata={
                'payment_id': payment.payment_id,
                'user_id': payment.user.id,
                'event_id': payment.invoice.event.id
            }
        )
        
        # Update payment with Stripe data
        payment.transaction_id = intent.id
        payment.gateway_response = {
            'client_secret': intent.client_secret,
            'payment_intent_id': intent.id
        }
        payment.save()
        
        return JsonResponse({
            'success': True,
            'client_secret': intent.client_secret,
            'payment_id': payment.payment_id
        })
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe error: {str(e)}")
        payment.status = 'failed'
        payment.gateway_response = {'error': str(e)}
        payment.save()
        
        return JsonResponse({
            'success': False,
            'error': 'Payment processing failed. Please try again.'
        })

def process_razorpay_payment(request, payment):
    """Process payment using Razorpay"""
    try:
        import razorpay
        
        # Initialize Razorpay client
        client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
        
        # Create Razorpay order
        order_data = {
            'amount': int(payment.amount * 100),  # Convert to paise
            'currency': payment.currency,
            'receipt': payment.payment_id,
            'notes': {
                'payment_id': payment.payment_id,
                'user_id': payment.user.id,
                'event_id': payment.invoice.event.id
            }
        }
        
        order = client.order.create(data=order_data)
        
        # Update payment with Razorpay data
        payment.transaction_id = order['id']
        payment.gateway_response = {
            'order_id': order['id'],
            'amount': order['amount'],
            'currency': order['currency']
        }
        payment.save()
        
        return JsonResponse({
            'success': True,
            'order_id': order['id'],
            'amount': order['amount'],
            'currency': order['currency'],
            'payment_id': payment.payment_id
        })
        
    except Exception as e:
        logger.error(f"Razorpay error: {str(e)}")
        payment.status = 'failed'
        payment.gateway_response = {'error': str(e)}
        payment.save()
        
        return JsonResponse({
            'success': False,
            'error': 'Payment processing failed. Please try again.'
        })

def process_other_payment(request, payment):
    """Process payment using other methods (bank transfer, etc.)"""
    try:
        # For bank transfers, mark as pending and send instructions
        payment.status = 'pending'
        payment.save()
        
        # Send payment instructions email
        send_payment_instructions_email(payment)
        
        return JsonResponse({
            'success': True,
            'payment_id': payment.payment_id,
            'message': 'Payment instructions sent to your email.'
        })
        
    except Exception as e:
        logger.error(f"Other payment error: {str(e)}")
        payment.status = 'failed'
        payment.gateway_response = {'error': str(e)}
        payment.save()
        
        return JsonResponse({
            'success': False,
            'error': 'Payment processing failed. Please try again.'
        })

@csrf_exempt
@require_POST
def stripe_webhook(request):
    """Handle Stripe webhooks"""
    payload = request.body
    sig_header = request.META.get('HTTP_STRIPE_SIGNATURE')
    
    try:
        event = stripe.Webhook.construct_event(
            payload, sig_header, settings.STRIPE_WEBHOOK_SECRET
        )
    except ValueError as e:
        logger.error(f"Invalid payload: {str(e)}")
        return HttpResponse(status=400)
    except stripe.error.SignatureVerificationError as e:
        logger.error(f"Invalid signature: {str(e)}")
        return HttpResponse(status=400)
    
    # Handle the event
    if event['type'] == 'payment_intent.succeeded':
        handle_payment_success(event['data']['object'])
    elif event['type'] == 'payment_intent.payment_failed':
        handle_payment_failure(event['data']['object'])
    elif event['type'] == 'charge.refunded':
        handle_refund_success(event['data']['object'])
    
    return HttpResponse(status=200)

def handle_payment_success(payment_intent):
    """Handle successful payment"""
    try:
        payment = Payment.objects.get(transaction_id=payment_intent['id'])
        payment.status = 'completed'
        payment.payment_date = timezone.now()
        payment.gateway_response['success'] = True
        payment.save()
        
        # Update invoice
        invoice = payment.invoice
        invoice.status = 'paid'
        invoice.paid_amount = payment.amount
        invoice.save()
        
        # Send confirmation email
        send_payment_confirmation_email(payment)
        
        # Generate and send invoice
        send_invoice_email(payment)
        
        logger.info(f"Payment {payment.payment_id} completed successfully")
        
    except Payment.DoesNotExist:
        logger.error(f"Payment not found for transaction {payment_intent['id']}")
    except Exception as e:
        logger.error(f"Error handling payment success: {str(e)}")

def handle_payment_failure(payment_intent):
    """Handle failed payment"""
    try:
        payment = Payment.objects.get(transaction_id=payment_intent['id'])
        payment.status = 'failed'
        payment.gateway_response['failure_reason'] = payment_intent.get('last_payment_error', {}).get('message', 'Unknown error')
        payment.save()
        
        # Send failure notification
        send_payment_failure_email(payment)
        
        logger.info(f"Payment {payment.payment_id} failed")
        
    except Payment.DoesNotExist:
        logger.error(f"Payment not found for transaction {payment_intent['id']}")
    except Exception as e:
        logger.error(f"Error handling payment failure: {str(e)}")

def handle_refund_success(charge):
    """Handle successful refund"""
    try:
        payment = Payment.objects.get(transaction_id=charge['payment_intent'])
        
        # Create refund record
        refund = Refund.objects.create(
            payment=payment,
            user=payment.user,
            amount=Decimal(charge['amount_refunded']) / 100,
            reason='Customer request',
            status='completed',
            transaction_id=charge['id'],
            refund_date=timezone.now()
        )
        
        # Update payment status
        payment.status = 'refunded'
        payment.save()
        
        # Send refund confirmation
        send_refund_confirmation_email(refund)
        
        logger.info(f"Refund {refund.refund_id} completed successfully")
        
    except Payment.DoesNotExist:
        logger.error(f"Payment not found for charge {charge['id']}")
    except Exception as e:
        logger.error(f"Error handling refund: {str(e)}")

@login_required
def payment_success(request):
    """Handle successful payment redirect"""
    payment_id = request.GET.get('payment_id')
    
    try:
        payment = Payment.objects.get(payment_id=payment_id, user=request.user)
        
        context = {
            'payment': payment,
            'event': payment.invoice.event,
        }
        
        if payment.status == 'completed':
            messages.success(request, 'Payment completed successfully!')
            return render(request, 'payments/payment_success.html', context)
        else:
            messages.warning(request, 'Payment is being processed. You will receive a confirmation email shortly.')
            return render(request, 'payments/payment_success.html', context)
            
    except Payment.DoesNotExist:
        messages.error(request, 'Payment not found.')
        return render(request, 'payments/payment_success.html', {'payment': None})

@login_required
def payment_cancel(request):
    """Handle cancelled payment"""
    messages.info(request, 'Payment was cancelled.')
    return redirect('events:event_list')

@login_required
def request_refund(request, payment_id):
    """Request refund for payment"""
    payment = get_object_or_404(Payment, payment_id=payment_id, user=request.user)
    
    if request.method == 'POST':
        reason = request.POST.get('reason')
        amount = Decimal(request.POST.get('amount', payment.amount))
        
        if amount > payment.amount:
            messages.error(request, 'Refund amount cannot exceed payment amount.')
            return redirect('payments:payment_list')
        
        # Create refund request
        refund = Refund.objects.create(
            payment=payment,
            user=request.user,
            amount=amount,
            reason=reason,
            status='pending'
        )
        
        # Process refund based on payment method
        if payment.payment_method.name == 'stripe':
            process_stripe_refund(refund)
        elif payment.payment_method.name == 'razorpay':
            process_razorpay_refund(refund)
        else:
            # Manual refund process
            refund.status = 'processing'
            refund.save()
            send_refund_request_email(refund)
        
        messages.success(request, 'Refund request submitted successfully.')
        return redirect('payments:payment_list')
    
    return render(request, 'payments/request_refund.html', {'payment': payment})

@login_required
def download_invoice(request, invoice_id):
    """Download PDF invoice"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Check if user has permission to download this invoice
    if not (request.user == invoice.user or request.user.user_type in ['admin', 'manager']):
        messages.error(request, 'You do not have permission to download this invoice.')
        return redirect('payments:payment_list')
    
    # Generate PDF invoice
    response = generate_pdf_invoice(invoice.event, request.user)
    
    if response:
        response['Content-Disposition'] = f'attachment; filename="invoice_{invoice.invoice_number}.pdf"'
        return response
    else:
        messages.error(request, 'Error generating invoice. Please try again.')
        return redirect('payments:payment_list')

@login_required
def invoice_detail(request, invoice_id):
    """Display invoice details"""
    invoice = get_object_or_404(Invoice, id=invoice_id)
    
    # Check if user has permission to view this invoice
    if not (request.user == invoice.user or request.user.user_type in ['admin', 'manager']):
        messages.error(request, 'You do not have permission to view this invoice.')
        return redirect('payments:payment_list')
    
    context = {
        'invoice': invoice,
        'payments': invoice.payments.all(),
        'refunds': Refund.objects.filter(payment__invoice=invoice)
    }
    
    return render(request, 'payments/invoice_detail.html', context)

def process_stripe_refund(refund):
    """Process refund through Stripe"""
    try:
        stripe_refund = stripe.Refund.create(
            payment_intent=refund.payment.transaction_id,
            amount=int(refund.amount * 100),
            reason='requested_by_customer'
        )
        
        refund.transaction_id = stripe_refund.id
        refund.status = 'processing'
        refund.save()
        
    except stripe.error.StripeError as e:
        logger.error(f"Stripe refund error: {str(e)}")
        refund.status = 'failed'
        refund.gateway_response = {'error': str(e)}
        refund.save()

def process_razorpay_refund(refund):
    """Process refund through Razorpay"""
    try:
        import razorpay
        
        client = razorpay.Client(auth=(settings.RAZORPAY_KEY_ID, settings.RAZORPAY_KEY_SECRET))
        
        razorpay_refund = client.payment.refund(
            refund.payment.transaction_id,
            {
                'amount': int(refund.amount * 100),
                'speed': 'normal'
            }
        )
        
        refund.transaction_id = razorpay_refund['id']
        refund.status = 'processing'
        refund.save()
        
    except Exception as e:
        logger.error(f"Razorpay refund error: {str(e)}")
        refund.status = 'failed'
        refund.gateway_response = {'error': str(e)}
        refund.save()

# Email functions
def send_payment_confirmation_email(payment):
    """Send payment confirmation email"""
    subject = f'Payment Confirmation - {payment.invoice.event.title}'
    context = {
        'payment': payment,
        'event': payment.invoice.event,
        'user': payment.user
    }
    
    html_message = render_to_string('payments/emails/payment_confirmation.html', context)
    plain_message = render_to_string('payments/emails/payment_confirmation.txt', context)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [payment.user.email],
        html_message=html_message
    )

def send_payment_instructions_email(payment):
    """Send payment instructions email"""
    subject = f'Payment Instructions - {payment.invoice.event.title}'
    context = {
        'payment': payment,
        'event': payment.invoice.event,
        'user': payment.user
    }
    
    html_message = render_to_string('payments/emails/payment_instructions.html', context)
    plain_message = render_to_string('payments/emails/payment_instructions.txt', context)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [payment.user.email],
        html_message=html_message
    )

def send_payment_failure_email(payment):
    """Send payment failure notification"""
    subject = f'Payment Failed - {payment.invoice.event.title}'
    context = {
        'payment': payment,
        'event': payment.invoice.event,
        'user': payment.user
    }
    
    html_message = render_to_string('payments/emails/payment_failure.html', context)
    plain_message = render_to_string('payments/emails/payment_failure.txt', context)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [payment.user.email],
        html_message=html_message
    )

def send_refund_confirmation_email(refund):
    """Send refund confirmation email"""
    subject = f'Refund Confirmation - {refund.payment.invoice.event.title}'
    context = {
        'refund': refund,
        'payment': refund.payment,
        'event': refund.payment.invoice.event,
        'user': refund.user
    }
    
    html_message = render_to_string('payments/emails/refund_confirmation.html', context)
    plain_message = render_to_string('payments/emails/refund_confirmation.txt', context)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [refund.user.email],
        html_message=html_message
    )

def send_refund_request_email(refund):
    """Send refund request notification to admin"""
    subject = f'Refund Request - {refund.payment.invoice.event.title}'
    context = {
        'refund': refund,
        'payment': refund.payment,
        'event': refund.payment.invoice.event,
        'user': refund.user
    }
    
    html_message = render_to_string('payments/emails/refund_request.html', context)
    plain_message = render_to_string('payments/emails/refund_request.txt', context)
    
    send_mail(
        subject,
        plain_message,
        settings.DEFAULT_FROM_EMAIL,
        [getattr(settings, 'ADMIN_EMAIL', 'admin@example.com')],
        html_message=html_message
    )

def send_invoice_email(payment):
    """Send invoice email to user"""
    try:
        subject = f'Invoice for {payment.invoice.event.title}'
        message = f"""
        Dear {payment.user.get_full_name()},
        
        Please find attached the invoice for your booking.
        
        Event: {payment.invoice.event.title}
        Amount: ₹{payment.amount}
        Date: {payment.invoice.event.start_date}
        
        Thank you for choosing our service!
        
        Best regards,
        Event Manager Team
        """
        
        # Send email with PDF attachment
        send_mail(
            subject,
            message,
            settings.DEFAULT_FROM_EMAIL,
            [payment.user.email],
            fail_silently=False,
        )
        
    except Exception as e:
        logger.error(f"Error sending invoice email: {e}")


@login_required
def export_payments(request):
    """Export payments data to Excel"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io
        
        # Get filter parameters
        status_filter = request.GET.get('status', '')
        method_filter = request.GET.get('method', '')
        date_filter = request.GET.get('date', '')
        amount_filter = request.GET.get('amount', '')
        search_filter = request.GET.get('search', '')
        
        # Build query
        payments = Payment.objects.filter(user=request.user)
        
        if status_filter:
            payments = payments.filter(status=status_filter)
        if method_filter:
            payments = payments.filter(payment_method=method_filter)
        if date_filter:
            payments = payments.filter(created_at__date=date_filter)
        if amount_filter:
            if amount_filter == 'low':
                payments = payments.filter(amount__lt=1000)
            elif amount_filter == 'medium':
                payments = payments.filter(amount__gte=1000, amount__lt=5000)
            elif amount_filter == 'high':
                payments = payments.filter(amount__gte=5000)
        if search_filter:
            payments = payments.filter(
                Q(payment_id__icontains=search_filter) |
                Q(invoice__event__title__icontains=search_filter)
            )
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Payments Report"
        
        # Add headers
        headers = [
            'Payment ID', 'Event', 'Amount', 'Status', 'Payment Method',
            'Created Date', 'Completed Date', 'Invoice Number'
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        for row, payment in enumerate(payments, 2):
            ws.cell(row=row, column=1, value=payment.payment_id)
            ws.cell(row=row, column=2, value=payment.invoice.event.title if payment.invoice else 'N/A')
            ws.cell(row=row, column=3, value=float(payment.amount))
            ws.cell(row=row, column=4, value=payment.status)
            ws.cell(row=row, column=5, value=payment.payment_method)
            ws.cell(row=row, column=6, value=payment.created_at.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=7, value=payment.completed_at.strftime('%Y-%m-%d %H:%M') if payment.completed_at else 'N/A')
            ws.cell(row=row, column=8, value=payment.invoice.invoice_number if payment.invoice else 'N/A')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payments_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'Error exporting payments: {str(e)}')
        return redirect('payments:payment_list')


def generate_paytm_qr_code(payment_amount, payment_id):
    """Generate Paytm QR code for payment"""
    # Create QR code data for Paytm - Replace with your actual Paytm number
    qr_data = f"upi://pay?pa=9725949961@paytm&pn=360° Event Manager&am={payment_amount}&tn=Payment for booking {payment_id}"
    
    # Generate QR code
    qr = qrcode.QRCode(version=1, box_size=10, border=5)
    qr.add_data(qr_data)
    qr.make(fit=True)
    
    # Create image
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to base64
    buffer = BytesIO()
    img.save(buffer, format='PNG')
    img_str = base64.b64encode(buffer.getvalue()).decode()
    
    return img_str

@login_required
def paytm_payment(request, payment_id):
    """Handle Paytm QR code payment"""
    if request.method == 'POST':
        try:
            payment = Payment.objects.get(id=payment_id, user=request.user)
            
            # Generate QR code
            qr_code = generate_paytm_qr_code(payment.amount, payment.id)
            
            # Set payment expiry (2 minutes)
            payment.expires_at = timezone.now() + timedelta(minutes=2)
            payment.save()
            
            context = {
                'payment': payment,
                'qr_code': qr_code,
                'expires_at': payment.expires_at,
                'payment_id': payment.id
            }
            return render(request, 'payments/paytm_payment.html', context)
            
        except Payment.DoesNotExist:
            messages.error(request, 'Payment not found')
            return redirect('payments:payment_list')
    
    return redirect('payments:payment_list')

@login_required
def paytm_payment_confirmation(request, payment_id):
    """Handle Paytm payment confirmation from admin"""
    if request.method == 'POST':
        try:
            payment = Payment.objects.get(id=payment_id)
            action = request.POST.get('action')
            
            if action == 'received':
                # Mark payment as successful
                payment.status = 'completed'
                payment.completed_at = timezone.now()
                payment.save()
                
                # Generate ticket and invoice
                generate_ticket_and_invoice(payment)
                
                messages.success(request, 'Payment confirmed successfully!')
                return JsonResponse({'success': True, 'message': 'Payment confirmed'})
                
            elif action == 'not_received':
                # Mark payment as failed
                payment.status = 'failed'
                payment.failed_at = timezone.now()
                payment.save()
                
                messages.error(request, 'Payment marked as not received')
                return JsonResponse({'success': True, 'message': 'Payment marked as failed'})
                
        except Payment.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Payment not found'})
    
    return JsonResponse({'success': False, 'error': 'Invalid request'})

def generate_ticket_and_invoice(payment):
    """Generate ticket and invoice for successful payment"""
    try:
        # Generate ticket
        ticket = generate_ticket(payment)
        
        # Generate invoice
        invoice = generate_invoice(payment)
        
        # Send confirmation email
        send_payment_confirmation_email(payment, ticket, invoice)
        
    except Exception as e:
        logger.error(f"Error generating ticket/invoice: {e}")

def generate_ticket(payment):
    """Generate ticket for the booking"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.blue,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.darkblue,
            spaceAfter=20
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=12
        )
        
        # Build PDF content
        story = []
        
        # Header
        story.append(Paragraph("🎫 EVENT TICKET", title_style))
        story.append(Spacer(1, 20))
        
        # Ticket details
        ticket_data = [
            ['Ticket Number:', f"TKT-{payment.id:06d}"],
            ['Event:', payment.invoice.event.title if payment.invoice else 'N/A'],
            ['Date:', payment.invoice.event.start_date.strftime('%B %d, %Y') if payment.invoice else 'N/A'],
            ['Time:', f"{payment.invoice.event.start_time} - {payment.invoice.event.end_time}" if payment.invoice else 'N/A'],
            ['Venue:', payment.invoice.event.venue.name if payment.invoice else 'N/A'],
            ['Attendee:', payment.user.get_full_name()],
            ['Email:', payment.user.email],
            ['Amount Paid:', f"₹{payment.amount}"],
            ['Payment Date:', payment.completed_at.strftime('%B %d, %Y %H:%M') if payment.completed_at else 'N/A'],
            ['Status:', 'CONFIRMED']
        ]
        
        # Create table
        ticket_table = Table(ticket_data, colWidths=[2*inch, 4*inch])
        ticket_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(ticket_table)
        story.append(Spacer(1, 30))
        
        # Terms and conditions
        story.append(Paragraph("Terms & Conditions:", subtitle_style))
        terms = [
            "• This ticket is non-transferable and non-refundable",
            "• Please arrive 15 minutes before the event",
            "• Valid ID proof may be required for entry",
            "• Photography may be restricted during the event",
            "• The organizer reserves the right to modify event details",
            "• In case of cancellation, refunds will be processed within 7-10 business days"
        ]
        
        for term in terms:
            story.append(Paragraph(term, normal_style))
        
        story.append(Spacer(1, 20))
        
        # Contact information
        story.append(Paragraph("Contact Information:", subtitle_style))
        contact_info = [
            "📞 Phone: +91 98765 43210",
            "📧 Email: info@360eventmanager.com",
            "🌐 Website: www.360eventmanager.com",
            "📍 Address: Ahmedabad, Gujarat"
        ]
        
        for info in contact_info:
            story.append(Paragraph(info, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error generating ticket PDF: {e}")
        return None

def generate_invoice(payment):
    """Generate invoice for the payment"""
    try:
        from reportlab.pdfgen import canvas
        from reportlab.lib.pagesizes import letter
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
        
        # Create PDF buffer
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        
        # Create custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.blue,
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        subtitle_style = ParagraphStyle(
            'CustomSubtitle',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.darkblue,
            spaceAfter=20
        )
        
        normal_style = ParagraphStyle(
            'CustomNormal',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=12
        )
        
        # Build PDF content
        story = []
        
        # Header
        story.append(Paragraph("📄 INVOICE", title_style))
        story.append(Spacer(1, 20))
        
        # Invoice details
        invoice_data = [
            ['Invoice Number:', f"INV-{payment.id:06d}"],
            ['Invoice Date:', payment.completed_at.strftime('%B %d, %Y') if payment.completed_at else 'N/A'],
            ['Due Date:', payment.completed_at.strftime('%B %d, %Y') if payment.completed_at else 'N/A'],
            ['Payment Status:', 'PAID'],
            ['Payment Method:', payment.payment_method],
            ['Customer Name:', payment.user.get_full_name()],
            ['Customer Email:', payment.user.email],
            ['Customer Phone:', payment.user.phone if hasattr(payment.user, 'phone') else 'N/A'],
        ]
        
        # Create table
        invoice_table = Table(invoice_data, colWidths=[2*inch, 4*inch])
        invoice_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(invoice_table)
        story.append(Spacer(1, 30))
        
        # Service details
        story.append(Paragraph("Service Details:", subtitle_style))
        
        service_data = [
            ['Description', 'Quantity', 'Unit Price', 'Amount'],
            [payment.invoice.event.title if payment.invoice else 'Event Booking', '1', f"₹{payment.amount}", f"₹{payment.amount}"],
        ]
        
        # Add venue cost if available
        if payment.invoice and payment.invoice.event.venue:
            service_data.append([
                f"Venue: {payment.invoice.event.venue.name}",
                '1',
                f"₹{payment.invoice.event.venue_cost}",
                f"₹{payment.invoice.event.venue_cost}"
            ])
        
        # Create service table
        service_table = Table(service_data, colWidths=[2.5*inch, 1*inch, 1.5*inch, 1.5*inch])
        service_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(service_table)
        story.append(Spacer(1, 20))
        
        # Total
        total_data = [
            ['Subtotal:', f"₹{payment.amount}"],
            ['Tax (18% GST):', f"₹{payment.amount * 0.18:.2f}"],
            ['Total Amount:', f"₹{payment.amount * 1.18:.2f}"]
        ]
        
        total_table = Table(total_data, colWidths=[4*inch, 2*inch])
        total_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, -1), (-1, -1), colors.blue),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
        ]))
        
        story.append(total_table)
        story.append(Spacer(1, 30))
        
        # Payment terms
        story.append(Paragraph("Payment Terms:", subtitle_style))
        terms = [
            "• Payment is due upon receipt of this invoice",
            "• Late payments may incur additional charges",
            "• All payments are non-refundable unless otherwise specified",
            "• For any queries, contact our support team"
        ]
        
        for term in terms:
            story.append(Paragraph(term, normal_style))
        
        story.append(Spacer(1, 20))
        
        # Company information
        story.append(Paragraph("Company Information:", subtitle_style))
        company_info = [
            "🏢 360° Event Manager",
            "📞 Phone: +91 98765 43210",
            "📧 Email: info@360eventmanager.com",
            "🌐 Website: www.360eventmanager.com",
            "📍 Address: Ahmedabad, Gujarat, India",
            "📄 GST Number: 24XXXXXXXXX1Z5"
        ]
        
        for info in company_info:
            story.append(Paragraph(info, normal_style))
        
        # Build PDF
        doc.build(story)
        buffer.seek(0)
        
        return buffer
        
    except Exception as e:
        logger.error(f"Error generating invoice PDF: {e}")
        return None
