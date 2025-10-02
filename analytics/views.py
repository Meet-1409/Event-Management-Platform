from django.shortcuts import render
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Avg, Q
from django.utils import timezone
from datetime import datetime, timedelta
import json

from users.models import CustomUser
from events.models import Event, Registration
from venues.models import Venue
from payments.models import Payment
from communications.models import Conversation, Message


@login_required
def analytics_dashboard(request):
    """Main analytics dashboard view"""
    return render(request, 'analytics/dashboard.html')


@login_required
def get_dashboard_data(request):
    """API endpoint to get real-time dashboard data"""
    try:
        # Get date range (last 6 months by default)
        end_date = timezone.now()
        start_date = end_date - timedelta(days=180)
        
        # User statistics
        total_users = CustomUser.objects.count()
        new_users_this_month = CustomUser.objects.filter(
            date_joined__gte=start_date
        ).count()
        
        # Event statistics
        total_events = Event.objects.count()
        active_events = Event.objects.filter(
            start_date__gte=timezone.now()
        ).count()
        
        # Booking statistics
        total_bookings = Registration.objects.count()
        confirmed_bookings = Registration.objects.count()  # All registrations are considered confirmed
        
        # Revenue statistics
        total_revenue = Payment.objects.filter(status='completed').aggregate(
            total=Sum('amount')
        )['total'] or 0
        
        monthly_revenue = Payment.objects.filter(
            status='completed',
            created_at__gte=start_date
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        # Venue statistics
        total_venues = Venue.objects.count()
        active_venues = Venue.objects.filter(is_active=True).count()
        
        # Communication statistics
        total_conversations = Conversation.objects.count()
        total_messages = Message.objects.count()
        
        # Monthly data for charts
        monthly_data = get_monthly_data(start_date, end_date)
        
        data = {
            'users': {
                'total': total_users,
                'new_this_month': new_users_this_month,
                'growth_rate': calculate_growth_rate('users', start_date, end_date)
            },
            'events': {
                'total': total_events,
                'active': active_events,
                'monthly_data': monthly_data['events']
            },
            'bookings': {
                'total': total_bookings,
                'confirmed': confirmed_bookings,
                'conversion_rate': calculate_conversion_rate(confirmed_bookings, total_bookings)
            },
            'revenue': {
                'total': float(total_revenue),
                'monthly': float(monthly_revenue),
                'monthly_data': monthly_data['revenue']
            },
            'venues': {
                'total': total_venues,
                'active': active_venues
            },
            'communications': {
                'conversations': total_conversations,
                'messages': total_messages
            }
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


def get_monthly_data(start_date, end_date):
    """Get monthly aggregated data for charts"""
    months = []
    events_data = []
    revenue_data = []
    
    current_date = start_date
    while current_date <= end_date:
        month_start = current_date.replace(day=1)
        month_end = (month_start + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Events in this month
        events_count = Event.objects.filter(
            created_at__gte=month_start,
            created_at__lte=month_end
        ).count()
        
        # Revenue in this month
        revenue = Payment.objects.filter(
            status='completed',
            created_at__gte=month_start,
            created_at__lte=month_end
        ).aggregate(total=Sum('amount'))['total'] or 0
        
        months.append(current_date.strftime('%b'))
        events_data.append(events_count)
        revenue_data.append(float(revenue))
        
        current_date = (current_date + timedelta(days=32)).replace(day=1)
    
    return {
        'events': {
            'labels': months,
            'data': events_data
        },
        'revenue': {
            'labels': months,
            'data': revenue_data
        }
    }


def calculate_growth_rate(model_type, start_date, end_date):
    """Calculate growth rate for a specific model"""
    try:
        if model_type == 'users':
            current_count = CustomUser.objects.filter(
                date_joined__gte=start_date
            ).count()
            previous_count = CustomUser.objects.filter(
                date_joined__gte=start_date - timedelta(days=30),
                date_joined__lt=start_date
            ).count()
        elif model_type == 'events':
            current_count = Event.objects.filter(
                created_at__gte=start_date
            ).count()
            previous_count = Event.objects.filter(
                created_at__gte=start_date - timedelta(days=30),
                created_at__lt=start_date
            ).count()
        else:
            return 0
        
        if previous_count == 0:
            return 100 if current_count > 0 else 0
        
        growth_rate = ((current_count - previous_count) / previous_count) * 100
        return round(growth_rate, 2)
        
    except Exception:
        return 0


def calculate_conversion_rate(confirmed, total):
    """Calculate conversion rate"""
    if total == 0:
        return 0
    return round((confirmed / total) * 100, 2)


@login_required
def get_event_analytics(request):
    """Get detailed event analytics"""
    try:
        # Event categories
        event_categories = Event.objects.values('event_type').annotate(
            count=Count('id')
        ).order_by('-count')
        
        # Top performing venues
        top_venues = Venue.objects.annotate(
            booking_count=Count('events')
        ).order_by('-booking_count')[:5]
        
        # Event status distribution
        event_status = Event.objects.values('status').annotate(
            count=Count('id')
        )
        
        data = {
            'categories': list(event_categories),
            'top_venues': [
                {
                    'name': venue.name,
                    'bookings': venue.booking_count,
                                    'revenue': float(Event.objects.filter(
                    venue=venue
                ).aggregate(total=Sum('total_cost'))['total'] or 0)
                }
                for venue in top_venues
            ],
            'status_distribution': list(event_status)
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def get_revenue_analytics(request):
    """Get detailed revenue analytics"""
    try:
        # Payment methods distribution
        payment_methods = Payment.objects.values('payment_method').annotate(
            count=Count('id'),
            total=Sum('amount')
        )
        
        # Monthly revenue trend
        monthly_revenue = Payment.objects.filter(
            status='completed'
        ).extra(
            select={'month': "strftime('%%Y-%%m', created_at)"}
        ).values('month').annotate(
            total=Sum('amount')
        ).order_by('month')
        
        # Average transaction value
        avg_transaction = Payment.objects.filter(
            status='completed'
        ).aggregate(avg=Avg('amount'))['avg'] or 0
        
        data = {
            'payment_methods': list(payment_methods),
            'monthly_revenue': list(monthly_revenue),
            'avg_transaction': float(avg_transaction)
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def export_analytics(request):
    """Export analytics data"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        import io
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Analytics Report"
        
        # Add headers
        headers = ['Metric', 'Value', 'Date']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Add data
        current_date = timezone.now().strftime('%Y-%m-%d')
        data = [
            ['Total Users', CustomUser.objects.count(), current_date],
            ['Total Events', Event.objects.count(), current_date],
            ['Total Bookings', Registration.objects.count(), current_date],
            ['Total Revenue', float(Payment.objects.filter(status='completed').aggregate(total=Sum('amount'))['total'] or 0), current_date],
            ['Active Venues', Venue.objects.filter(is_active=True).count(), current_date],
        ]
        
        for row, (metric, value, date) in enumerate(data, 2):
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=date)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        from django.http import HttpResponse
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="analytics_report_{current_date}.xlsx"'
        
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
